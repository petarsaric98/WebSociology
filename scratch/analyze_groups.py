import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('pitanja matur.xlsx')
ws = wb.active

# Show more samples of Skupina zadataka and Zadatci povezivanja
count = 0
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    pitanje, tip, ponudeni, tocan, godina = row
    if tip == 'Skupina zadataka' and count < 5:
        count += 1
        print(f"\n=== Skupina {count} ===")
        print(f"  Pitanje: {str(pitanje)[:300]}")
        print(f"  Ponudeni: {str(ponudeni)[:300]}")
        print(f"  Tocan: {str(tocan)[:300]}")

# Check for multi-answer patterns in visestrukih kombinacija
print("\n\n=== Visestrukih kombinacija - odgovori ===")
count = 0
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    pitanje, tip, ponudeni, tocan, godina = row
    if tip == 'Zadatci višestrukih kombinacija' and count < 5:
        count += 1
        print(f"\n--- {count} ---")
        print(f"  Ponudeni: {ponudeni}")
        print(f"  Tocan: {tocan}")

# Check Zadatci povezivanja i sredivanja
print("\n\n=== Povezivanje i sređivanje ===")
count = 0
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    pitanje, tip, ponudeni, tocan, godina = row
    if tip == 'Zadatci povezivanja i sređivanja' and count < 5:
        count += 1
        print(f"\n--- {count} ---")
        print(f"  Pitanje: {str(pitanje)[:200]}")
        print(f"  Ponudeni: {str(ponudeni)[:400]}")
        print(f"  Tocan: {str(tocan)[:200]}")

# Check Zadatci dopunjavanja
print("\n\n=== Dopunjavanje ===")
count = 0
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    pitanje, tip, ponudeni, tocan, godina = row
    if tip == 'Zadatci dopunjavanja' and count < 5:
        count += 1
        print(f"\n--- {count} ---")
        print(f"  Pitanje: {str(pitanje)[:300]}")
        print(f"  Tocan: {str(tocan)[:200]}")

# Kratki odgovor samples
print("\n\n=== Kratki odgovor (5 samples) ===")
count = 0
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    pitanje, tip, ponudeni, tocan, godina = row
    if tip == 'Zadatci kratkog odgovora' and count < 5:
        count += 1
        print(f"\n--- {count} ---")
        print(f"  Pitanje: {str(pitanje)[:200]}")
        print(f"  Tocan: {str(tocan)[:200]}")

# Produzeni odgovor samples
print("\n\n=== Produzeni odgovor (3 samples) ===")
count = 0
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    pitanje, tip, ponudeni, tocan, godina = row
    if tip == 'Zadatci produženog odgovora' and count < 5:
        count += 1
        print(f"\n--- {count} ---")
        print(f"  Pitanje: {str(pitanje)[:300]}")
        print(f"  Tocan: {str(tocan)[:400]}")
