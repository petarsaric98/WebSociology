import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('pitanja matur.xlsx')
ws = wb.active

# Count question types
types = {}
years = {}
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    pitanje, tip, ponudeni, tocan, godina = row
    if tip:
        types[tip] = types.get(tip, 0) + 1
    if godina:
        years[godina] = years.get(godina, 0) + 1

print("=== TIPOVI PITANJA ===")
for k, v in sorted(types.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v}")

print("\n=== GODINE ===")
for k, v in sorted(years.items()):
    print(f"  {k}: {v}")

# Show sample questions for each type
print("\n=== UZORCI PO TIPU ===")
seen_types = set()
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    pitanje, tip, ponudeni, tocan, godina = row
    if tip and tip not in seen_types:
        seen_types.add(tip)
        print(f"\n--- Tip: {tip} ---")
        print(f"  Pitanje: {str(pitanje)[:200]}")
        print(f"  Ponudeni: {str(ponudeni)[:200]}")
        print(f"  Tocan: {str(tocan)[:200]}")
        print(f"  Godina: {godina}")
